from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
import logging

def generar_archivo_excel(datos_db, fecha_ayer_1, itera_nomzona, path_arch_audi1, path_carp_email):
    # Suponiendo que 'data1' es una lista de listas con los datos
    data1 = datos_db # Datos para la primera hoja traidos desde variable de Rocketbot

    # Crear un nuevo libro de Excel
    wb = Workbook()

    ########## DATA 1 ##########

    # Renombrar la primera hoja
    ws1 = wb.active
    ws1.title = "Hoja1"

    # Agregar los encabezados a la primera hoja
    header_row = ["FECHA", "CÓDIGO ZONA", "ZONA", "C.COSTO", "CÓDIGO PV", "NOMBRE PVT", "IDENTIFICACIÓN", "SALDO"]
    ws1.append(header_row)

    # Agregar los datos a las celdas de la primera hoja
    for row in data1:
        ws1.append(row)

    # Aplicar formato a los encabezados
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="07C700", end_color="07C700", fill_type="solid")
    for cell in ws1["1:1"]:
        cell.font = header_font
        cell.fill = header_fill

    # Agregar bordes a la tabla
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row, min_col=1, max_col=ws1.max_column):
        for cell in row:
            cell.border = thin_border

    # Obtener el ancho máximo de cada columna
    column_widths = []
    for col in ws1.columns:
        max_length = 0
        column_letter = col[0].column_letter  # Obtener las letras de la columna
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        column_widths.append((column_letter, adjusted_width))

    # Establecer los anchos de columna
    for col, width in column_widths:
        ws1.column_dimensions[col].width = width


    # Guardar el libro de Excel en ruta dependiendo la condición
    wb.save(f"{path_arch_audi1}_{fecha_ayer_1}_{itera_nomzona}.xlsx")
    wb.save(f"{path_carp_email}_{fecha_ayer_1}_{itera_nomzona}.xlsx")
    logging.info(f'Archivo: {path_carp_email}_{fecha_ayer_1}_{itera_nomzona}.xlsx creado.')

def generar_archivo_excel_2(datos_db_2, fecha_ayer_1, itera_nom_mont, path_arch_audi1, path_carp_email):
    # Suponiendo que 'data1' es una lista de listas con los datos
    data1 = datos_db_2 # Datos para la primera hoja traidos desde variable de Rocketbot

    # Crear un nuevo libro de Excel
    wb = Workbook()

    ########## DATA 1 ##########

    # Renombrar la primera hoja
    ws1 = wb.active
    ws1.title = "Hoja1"

    # Agregar los encabezados a la primera hoja
    header_row = ["FECHA", "CÓDIGO ZONA", "ZONA", "C.COSTO", "CÓDIGO PV", "NOMBRE PVT", "IDENTIFICACIÓN", "SALDO"]
    ws1.append(header_row)

    # Agregar los datos a las celdas de la primera hoja
    for row in data1:
        ws1.append(row)

    # Aplicar formato a los encabezados
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="07C700", end_color="07C700", fill_type="solid")
    for cell in ws1["1:1"]:
        cell.font = header_font
        cell.fill = header_fill

    # Agregar bordes a la tabla
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row, min_col=1, max_col=ws1.max_column):
        for cell in row:
            cell.border = thin_border

    # Obtener el ancho máximo de cada columna
    column_widths = []
    for col in ws1.columns:
        max_length = 0
        column_letter = col[0].column_letter  # Obtener las letras de la columna
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        column_widths.append((column_letter, adjusted_width))

    # Establecer los anchos de columna
    for col, width in column_widths:
        ws1.column_dimensions[col].width = width

    # Guardar el libro de Excel en ruta dependiendo la condición

    wb.save(f"{path_arch_audi1}_{fecha_ayer_1}_{itera_nom_mont}.xlsx")
    wb.save(f"{path_carp_email}_{fecha_ayer_1}_{itera_nom_mont}.xlsx")
    logging.info(f'Archivo: {path_carp_email}_{fecha_ayer_1}_{itera_nom_mont}.xlsx creado.')

